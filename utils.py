# utils.py
import io
import json
import re
import time
import warnings
import datetime as dt
from threading import Lock
from typing import Optional, Tuple, Dict

import pandas as pd
import requests
from isoweek import Week
from openpyxl import load_workbook

from config import settings


warnings.filterwarnings("ignore")

# сюда складываются МО, которые упали в процессе выгрузки талонов
cancelLPU = pd.DataFrame()

# (если нужен общий лок на операции с i38/шаблоном — оставил)
_I38_LOCK = Lock()


# -----------------------------
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# -----------------------------
def get_prev_week_sheet_name(today: Optional[dt.date] = None) -> str:
    """
    Возвращает строку 'YYYY-MM-DD - YYYY-MM-DD' для прошлой ISO-недели.
    Пример: если сегодня 2026-02-17, вернёт '2026-02-09 - 2026-02-15'
    """
    today = today or dt.date.today()
    iso = today.isocalendar()
    year = iso[0]
    week = iso[1]

    if week == 1:
        prev_year = year - 1
        prev_week = dt.date(prev_year, 12, 28).isocalendar()[1]  # последняя ISO-неделя года
    else:
        prev_year = year
        prev_week = week - 1

    monday = dt.date.fromisocalendar(prev_year, prev_week, 1)
    sunday = dt.date.fromisocalendar(prev_year, prev_week, 7)
    return f"{monday:%Y-%m-%d} - {sunday:%Y-%m-%d}"


def _clean_name(s: str) -> str:
    if s is None:
        return ""
    return "".join(filter(str.isalnum, str(s))).lower()


def _pick_col(df: pd.DataFrame, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def nz(v):
    return 0 if v is None else v


def ensure_formula(ws, addr, formula):
    v = ws[addr].value
    if not (isinstance(v, str) and v.startswith("=")):
        ws[addr].value = formula


def norm_id(v):
    if v is None:
        return None
    # Excel часто отдаёт числа как float
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_lpu(path: str):
    """
    Возвращает:
      df, clean_names(list), codes(list), name_col(str), code_col(str|None)
    """
    df = pd.read_excel(path)
    df.columns = (
        df.columns.astype(str)
        .str.replace("\u00a0", " ", regex=False)  # NBSP -> space
        .str.strip()
    )

    name_col = _pick_col(df, ["name", "Наменование МО", "Наименование МО", "МО", "NAME"])
    code_col = _pick_col(df, ["РМИС ID", "Код МО", "RMIS ID", "Код", "CODE"])

    if not name_col:
        raise ValueError(f"В {path} не найдена колонка с названием МО. Колонки: {list(df.columns)}")

    clean_names = [_clean_name(df.at[i, name_col]) for i in range(len(df))]
    codes = [norm_id(df.at[i, code_col]) if code_col else None for i in range(len(df))]

    return df, clean_names, codes, name_col, code_col


# -----------------------------
# ТАЛОНЫ (miac_table.xlsx) -> словари для Q/R/S
# -----------------------------
def load_ticket_maps(ticket_xlsx_path: str):
    """
    Возвращает 2 словаря:
      by_code:  RMIS_ID -> (portal_all, total, total_all)
      by_name:  cleaned_name -> (portal_all, total, total_all)

    Ожидается лист "Лист 1" и колонки:
      A: название МО
      B: РМИС ID
      C: PORTAL_ALL
      D: TOTAL
      E: TOTAL_ALL
    """
    wb_t = load_workbook(ticket_xlsx_path, data_only=True)
    ws_t = wb_t["Лист 1"]

    by_code: Dict[str, Tuple[float, float, float]] = {}
    by_name: Dict[str, Tuple[float, float, float]] = {}

    for r in range(1, ws_t.max_row + 1):
        code = norm_id(ws_t[f"B{r}"].value)
        name_clean = _clean_name(ws_t[f"A{r}"].value)

        portal_all = ws_t[f"C{r}"].value
        total = ws_t[f"D{r}"].value
        total_all = ws_t[f"E{r}"].value

        if code:
            by_code[code] = (portal_all, total, total_all)
        if name_clean:
            by_name[name_clean] = (portal_all, total, total_all)

    wb_t.close()
    return by_code, by_name


# -----------------------------
# ФИНАЛЬНЫЙ ОТЧЁТ: D/F/I/K/M/O по LPU2.xlsx
# -----------------------------
def build_final_excel_from_parse_bytes(
    parse_bytes: bytes,
    lpu_path: str,               # сюда передаём LPU2.xlsx (для строк/названий D/F/I/K/M/O)
    template_xlsx_path: str,     # settings.excel_path (xlsx с листом "Лист-шаблон")
    output_xlsx_path: str,
    sheet_name: str = "Лист-шаблон",
):
    # 1) читаем parse.json
    df = pd.read_json(io.BytesIO(parse_bytes))
    df = df.dropna().reset_index()  # как было (без drop=True)

    # 2) собираем карту метрик из parse.json по очищенному названию МО
    def getv(col_idx, row_idx):
        try:
            v = df.at[row_idx, col_idx]
            # ожидаемый формат: ["...", [число]]
            if isinstance(v, (list, tuple)) and len(v) >= 2:
                vv = v[1]
                if isinstance(vv, (list, tuple)) and len(vv) > 0:
                    return nz(vv[0])
            return 0
        except Exception:
            return 0

    parse_map = {}  # clean_name -> dict метрик
    for i in range(len(df)):
        nm = _clean_name(df.at[i, 0])
        if not nm:
            continue

        chern = getv(1, i)
        osh_all = getv(2, i)
        org_osh = getv(3, i)
        teh_osh = getv(4, i)
        usp = getv(5, i)
        fed = getv(6, i)

        # если вдруг дубли — суммируем
        cur = parse_map.get(nm)
        if cur is None:
            parse_map[nm] = {
                "chern": chern,
                "usp": usp,
                "osh_all": osh_all,
                "org_osh": org_osh,
                "teh_osh": teh_osh,
                "fed": fed,
            }
        else:
            cur["chern"] += chern
            cur["usp"] += usp
            cur["osh_all"] += osh_all
            cur["org_osh"] += org_osh
            cur["teh_osh"] += teh_osh
            cur["fed"] += fed

    # 3) читаем LPU2 (порядок/названия строк финального отчёта)
    lpu2_df, lpu2_clean, lpu2_codes, _, _ = _read_lpu(lpu_path)

    # 4) страховка: читаем LPU.xlsx (старые названия) -> map code -> old_clean_name
    #    нужно, если parse.json совпадает по названиям со старым LPU.xlsx,
    #    а в LPU2 названия другие
    src_by_code = {}
    try:
        _, lpu1_clean, lpu1_codes, _, _ = _read_lpu(settings.lpu_path)
        for i in range(len(lpu1_codes)):
            if lpu1_codes[i]:
                src_by_code[lpu1_codes[i]] = lpu1_clean[i]
    except Exception:
        src_by_code = {}

    # 5) читаем талоны (Q/R/S) из miac_table.xlsx
    ticket_by_code, ticket_by_name = load_ticket_maps(settings.excel_path_ticket)

    # 6) создаём новый лист прошлой недели в шаблоне
    wb = load_workbook(template_xlsx_path)
    week_sheet = get_prev_week_sheet_name()

    if week_sheet in wb.sheetnames:
        del wb[week_sheet]

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист-шаблон '{sheet_name}' не найден. Есть: {wb.sheetnames}")

    base_ws = wb[sheet_name]
    ws = wb.copy_worksheet(base_ws)
    ws.title = week_sheet
    ws["C1"] = week_sheet
    ws.sheet_state = "visible"

    # оставляем только новый лист (как у тебя было)
    keep_ws = ws
    for sh in wb.worksheets[:]:
        if sh is not keep_ws:
            wb.remove(sh)

    # попытка сделать активным (не критично)
    try:
        wb._sheets.remove(ws)
        wb._sheets.insert(2, ws)
        wb.active = 2
    except Exception:
        pass

    # 7) заполнение строк по LPU2 (D/F/I/K/M/O) + Q/R/S по miac_table
    limit = min(90, len(lpu2_clean))

    for i in range(0, limit):
        row = i + 4

        # метрики сначала пробуем по названию из LPU2
        m = parse_map.get(lpu2_clean[i])

        # если не нашли — пробуем через РМИС ID: LPU2(code) -> LPU1(old_name) -> parse_map(old_name)
        if m is None:
            code = lpu2_codes[i]
            old_name = src_by_code.get(code) if code else None
            if old_name:
                m = parse_map.get(old_name)

        if m:
            # D/F/I/K/M/O из parse.json
            ws[f"D{row}"] = nz(m["chern"])
            ws[f"F{row}"] = nz(m["usp"])
            ws[f"I{row}"] = nz(m["osh_all"])
            ws[f"K{row}"] = nz(m["org_osh"])
            ws[f"M{row}"] = nz(m["teh_osh"])
            ws[f"O{row}"] = nz(m["fed"])

            # C = D + F + I
            d_val = nz(ws[f"D{row}"].value)
            f_val = nz(ws[f"F{row}"].value)
            i_val = nz(ws[f"I{row}"].value)
            ws[f"C{row}"] = d_val + f_val + i_val

            # проценты
            ensure_formula(ws, f"E{row}", f"=D{row}/C{row}")
            ensure_formula(ws, f"G{row}", f"=F{row}/C{row}")
            ws[f"H{row}"].value = f"=IF(F{row}<>0,F{row}/(F{row}+K{row}+M{row}),0)"
            ensure_formula(ws, f"J{row}", f"=I{row}/C{row}")
            ensure_formula(ws, f"L{row}", f"=K{row}/C{row}")
            ensure_formula(ws, f"N{row}", f"=M{row}/C{row}")
            ensure_formula(ws, f"P{row}", f"=O{row}/C{row}")

        # Q/R/S — лучше по коду из LPU2 (стабильнее), иначе по названию
        code_key = norm_id(lpu2_codes[i])
        vals = ticket_by_code.get(code_key) if code_key else None
        if vals is None:
            vals = ticket_by_name.get(lpu2_clean[i])

        if vals:
            portal_all, total, total_all = vals
            ws[f"Q{row}"] = nz(portal_all)
            ws[f"R{row}"] = nz(total)
            ws[f"S{row}"] = nz(total_all)

        ensure_formula(ws, f"T{row}", f"=Q{row}/R{row}")

    wb.save(output_xlsx_path)
    wb.close()

    return output_xlsx_path


# -----------------------------
# ДАТЫ ПРОШЛОЙ НЕДЕЛИ (для выгрузки талонов)
# -----------------------------
def get_last_week_dates():
    d = dt.datetime.today()
    iso = d.isocalendar()
    # iso[1] = номер недели
    return (
        Week(d.year, (iso[1] - 1)).monday().strftime("%d.%m.%Y"),
        Week(d.year, (iso[1] - 1)).sunday().strftime("%d.%m.%Y"),
    )


# -----------------------------
# ВЫГРУЗКА ТАЛОНОВ (miac_table.xlsx) — КАК БЫЛО
# -----------------------------
def process_lpu(lpu_df: pd.DataFrame, ws, wb, cancel_lpu: pd.DataFrame = None):
    BASE_URL = settings.base_url
    LOGIN = settings.login
    PASSWORD = settings.password
    REPORT_ID = settings.report_id
    EMPLOYER = settings.employer
    DATE_FROM, DATE_TO = get_last_week_dates()

    session = requests.Session()

    # 0) Получение SYS_CACHE_UID
    url_js = f"{BASE_URL}/d3/~d3api"
    resp_js = session.get(url_js)
    resp_js.raise_for_status()

    match = re.search(r'D3Api\.SYS_CACHE_UID\s*=\s*"([a-f0-9]+)"', resp_js.text)
    if not match:
        raise ValueError("Не удалось найти SYS_CACHE_UID")
    SYS_CACHE_UID = match.group(1)
    print(f"[INFO] SYS_CACHE_UID = {SYS_CACHE_UID}")

    # 1) Авторизация
    url_login = f"{BASE_URL}/d3/request.php"
    params_login = {
        "cache_enabled": "1",
        "session_cache": "1",
        "modal": "1",
        "Form": "System/login",
        "cache": SYS_CACHE_UID,
    }
    data_login = {
        "request": json.dumps(
            {
                "Authorization": {
                    "type": "Module",
                    "params": {"DBPassword": PASSWORD, "DBLogin": LOGIN},
                }
            }
        )
    }
    resp_login = session.post(url_login, params=params_login, data=data_login)
    resp_login.raise_for_status()
    print("[INFO] Авторизация прошла успешно")

    # Bearer token
    login_json = resp_login.json()
    bearer_token = login_json.get("Authorization", {}).get("data", {}).get("access_token")
    if bearer_token:
        session.headers.update({"Authorization": f"Bearer {bearer_token}"})
        print("[INFO] Bearer token установлен")
    else:
        print("[WARN] Bearer токен не найден в ответе")

    # 2) Выбор LPU и получение данных
    for i in range(len(lpu_df)):
        LPU_ROW = lpu_df.iloc[i]

        url_lpu = f"{BASE_URL}/d3/request.php"
        params_lpu = {
            "cache_enabled": "1",
            "session_cache": "1",
            "modal": "1",
            "Form": "System/lpu",
            "cache": SYS_CACHE_UID,
        }
        data_lpu = {
            "request": json.dumps(
                {
                    "Authorization": {
                        "type": "Module",
                        "params": {
                            "LPU": str(LPU_ROW["РМИС ID"]),
                            "CABLAB": "",
                            "EMPLOYER": EMPLOYER,
                        },
                    }
                }
            )
        }
        resp_lpu = session.post(url_lpu, params=params_lpu, data=data_lpu)
        resp_lpu.raise_for_status()
        print(f"[INFO] LPU установлено {LPU_ROW['РМИС ID']}")

        # 3) Получение FormCache
        url_report_form = f"{BASE_URL}/getform.php"
        params_report_form = {
            "Form": "Reports/run",
            "modulename": "7465547992_report",
            "_rep_code": "lpu_talons_pmsp",
            "_exp_id": 0,
            "_fme_doc_id": 0,
            "modal": 1,
            "theme": "bars",
            "cache": SYS_CACHE_UID,
            "cache_enabled": 1,
            "session_cache": 1,
        }
        resp_form = session.post(url_report_form, params=params_report_form)
        resp_form.raise_for_status()

        form_cache = resp_form.headers.get("FormCache")
        if not form_cache:
            raise ValueError("Не удалось получить FormCache")
        print(f"[INFO] FormCache = {form_cache}")

        # 4) Получение данных отчёта (XML)
        url_multidata = f"{BASE_URL}/getmultidata.php"
        params_multidata = {
            "Form": "Reports/Statistic/lpu_talons",
            "reportid": REPORT_ID,
            "nooverflow": "true",
            "theme": "bars",
            "cache": SYS_CACHE_UID,
            "FormCache": form_cache,
        }
        data_multidata = {
            "_sysid": "1",
            "ds0[Form]": "Reports/Statistic/lpu_talons",
            "ds0[nooverflow]": "true",
            "ds0[DataSet]": "DS_TALONS",
            "ds0[DATE_FROM_g0]": DATE_FROM,
            "ds0[DATE_TO_g1]": DATE_TO,
            "ds0[CMP_REG_TYPE_g2]": 4,
            "ds0[EXCLUDE_TIME_TYPE_g4]": "2;6;27;39;41;96;116;136;149;420;421;422;423;424;74;77;155;400;401;402;403;75;17",
            "ds0[INCLUDE_SPECS_g6]": "1177;70;84;88;159;160;161;162;1179;61;123;117;1281;1282;1291;1292;1293;1294;1295;1300;5003;2155;2154;2153;9015;2489;2493;2499;2504;2511;2854;3104;3117;3127;3065;3077;3079;2501;2858;2860;2991;3096;3100;3111;3115;3124;3125;3132;3078;9014;2851;2861;3093;3105;3108;2498;2510;2815;2859;999;2976;2992;3106;2490;2492;2512;2977;3107;3139;2502;2509;19;2944;2978;3123;3126;3086;3075;2520;2855;2979;3094;3095;3099;3074;3076;78;5004;1372;87;2500;125;90;133;1056;5000;5001;203;2011;201;2103;2104;2515;2513",
            "ds0[_sysid]": "0",
        }

        try:
            resp_data = session.post(url_multidata, params=params_multidata, data=data_multidata)
            resp_data.raise_for_status()

            xml_clean = (
                resp_data.text
                .replace('<DataSet name="DS_TALONS" sysid="0">', "")
                .replace("</DataSet>", "")
            )

            df_xml = pd.read_xml(io.StringIO(xml_clean))

            print(f"[DEBUG] DataFrame shape: {df_xml.shape}")
            print(f"[DEBUG] Columns: {df_xml.columns.tolist()}")

            for k in range(3, 94):
                if lpu_df["РМИС ID"][i] == ws[f"B{k}"].value:
                    row_index = i + 4
                    ws[f"C{row_index}"] = df_xml["PORTAL_ALL"].sum()
                    ws[f"D{row_index}"] = df_xml["TOTAL"].sum()
                    ws[f"E{row_index}"] = df_xml["TOTAL_ALL"].sum()
                    ws[f"F{row_index}"] = dt.datetime.today()
                    break

            wb.save(settings.excel_path_ticket)
            print(f"[INFO] Сохранено для LPU {lpu_df['РМИС ID'][i]}")

        except Exception as e:
            print(f"[ERROR] Ошибка для LPU {lpu_df['РМИС ID'][i]}: {e}")
            if cancel_lpu is None:
                cancel_lpu = pd.DataFrame()

            new_row = pd.DataFrame(
                {
                    "Наменование МО": [lpu_df["Наменование МО"][i]],
                    "РМИС ID": [lpu_df["РМИС ID"][i]],
                }
            )
            global cancelLPU
            cancelLPU = pd.concat([cancelLPU, new_row], ignore_index=True)


def main_process():
    global cancelLPU

    LPU = pd.read_excel(settings.lpu_path)
    LPU.columns = (
        LPU.columns.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
    )

    if "РМИС ID" not in LPU.columns:
        for alt in ("Код МО", "RMIS ID", "Код", "CODE"):
            if alt in LPU.columns:
                LPU = LPU.rename(columns={alt: "РМИС ID"})
                break

    wb_obj = load_workbook(settings.excel_path_ticket)
    ws = wb_obj["Лист 1"]

    start_time = time.time()

    if not cancelLPU.empty:
        print("[INFO] Обработка неудачных LPU...")
        process_lpu(cancelLPU.drop_duplicates(), ws, wb_obj, cancelLPU)

    print("[INFO] Основная обработка...")
    process_lpu(LPU, ws, wb_obj, cancelLPU)

    wb_obj.save(settings.excel_path_ticket)
    wb_obj.close()

    duration = time.time() - start_time
    print(f"[INFO] Обработка завершена за {duration:.2f} сек")
    return duration