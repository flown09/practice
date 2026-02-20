import time
import os
from datetime import datetime
from fastapi import FastAPI, BackgroundTasks, HTTPException, Form, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from utils import main_process, get_last_week_dates
from config import settings
from pathlib import Path
import zipfile
import tempfile
from fastapi.responses import FileResponse
import uuid
from fastapi import UploadFile, File
from fastapi.concurrency import run_in_threadpool
from pathlib import Path
from utils import build_final_excel_from_parse_bytes
from pathlib import Path
from fastapi import UploadFile, File
import io
from openpyxl import load_workbook
from threading import Lock


app = FastAPI(title="MIAC Report Service")
templates = Jinja2Templates(directory="templates")

# Убираем глобальный scheduler, делаем его локальным
schedule_config = {
    "enabled": False,
    "day_of_week": 0,
    "hour": 10,
    "minute": 0
}

scheduler = None  # Глобальная переменная для текущего планировщика

REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)
TEMPLATE_LOCK = Lock()

@app.get("/template-info")
async def template_info():
    path = Path(settings.excel_path)
    if not path.exists():
        return {"exists": False}

    stat = path.stat()
    # попробуем прочитать названия листов (не обязательно, но полезно)
    sheetnames = None
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        sheetnames = wb.sheetnames
        wb.close()
    except Exception:
        sheetnames = None

    return {
        "exists": True,
        "filename": path.name,
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%d.%m.%Y %H:%M:%S"),
        "sheetnames": sheetnames,
    }


@app.get("/download-template")
async def download_template():
    path = Path(settings.excel_path)
    if not path.exists():
        raise HTTPException(404, "Шаблон не найден")
    return FileResponse(
        path=str(path),
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/upload-template")
async def upload_template(file: UploadFile = File(...)):
    # 1) проверка расширения
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Нужен Excel файл .xlsx")

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Пустой файл")

    # 2) валидация: файл реально открывается и содержит нужный лист
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        raise HTTPException(400, f"Не удалось открыть .xlsx: {e}")

    required_sheet = "Лист-шаблон"   # у тебя build_final_excel_from_parse_bytes ожидает это имя
    if required_sheet not in sheets:
        raise HTTPException(400, f"В шаблоне нет листа '{required_sheet}'. Есть: {sheets}")

    # 3) атомарная замена файла шаблона (без битых файлов при сбое)
    dst = Path(settings.excel_path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    tmp = dst.with_suffix(dst.suffix + ".tmp")

    with TEMPLATE_LOCK:
        tmp.write_bytes(raw)
        os.replace(str(tmp), str(dst))

    stat = dst.stat()
    return {
        "status": "ok",
        "filename": dst.name,
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%d.%m.%Y %H:%M:%S"),
        "sheetnames": sheets,
    }

@app.post("/upload-parse")
async def upload_parse(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".json"):
        raise HTTPException(400, "Нужен JSON файл (parse.json)")

    parse_bytes = await file.read()
    upload_id = uuid.uuid4().hex

    # куда сохраняем финальный файл
    out_path = REPORTS_DIR / f"final_{upload_id}.xlsx"

    # важно: обработка не в async потоке
    try:
        await run_in_threadpool(
            build_final_excel_from_parse_bytes,
            parse_bytes,
            settings.lpu_path,        # lpu.xlsx на сервере
            settings.excel_path,      # шаблон xlsx на сервере (положи сюда "и38 таблица МИАЦ.xlsx" или сделай отдельный setting)
            str(out_path),
            "Лист-шаблон"
        )
    except Exception as e:
        raise HTTPException(500, f"Ошибка обработки parse.json: {e}")

    return {
        "status": "ok",
        "upload_id": upload_id,
        "download_url": f"/download-final/{upload_id}"
    }

@app.post("/upload-parse-raw")
async def upload_parse_raw(file: UploadFile = File(...)):
    raw = await file.read()

    REPORTS_DIR = Path("reports")
    REPORTS_DIR.mkdir(exist_ok=True)

    save_path = REPORTS_DIR / "last_parse.json"
    save_path.write_bytes(raw)

    print(f"[upload-parse-raw] got {len(raw)} bytes -> {save_path}")

    return {"status": "ok", "bytes": len(raw), "filename": file.filename}


@app.get("/download-final/{upload_id}")
async def download_final(upload_id: str):
    path = REPORTS_DIR / f"final_{upload_id}.xlsx"
    if not path.exists():
        raise HTTPException(404, "Файл не найден")

    return FileResponse(
        path=str(path),
        filename="Отчет.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




def create_scheduler():
    """Создает и возвращает новый чистый планировщик"""
    global scheduler
    return BackgroundScheduler()

def scheduled_report():
    """Автоматическая выгрузка по расписанию"""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Автовыгрузка запущена")
    try:
        duration = main_process()
        print(f"[{(datetime.now().strftime('%H:%M:%S'))}] Автовыгрузка завершена за {duration:.2f}с")
    except Exception as e:
        print(f"[{(datetime.now().strftime('%H:%M:%S'))}] Ошибка автозагрузки: {e}")

def run_miac_direct():
    """Ручная выгрузка"""
    print(f"[{(datetime.now().strftime('%H:%M:%S'))}] Ручная выгрузка запущена")
    try:
        duration = main_process()
        print(f"[{(datetime.now().strftime('%H:%M:%S'))}] Обработка завершена за {duration:.2f} секунд")
    except Exception as e:
        print(f"[{(datetime.now().strftime('%H:%M:%S'))}] Ошибка обработки: {e}")

@app.on_event("startup")
async def startup():
    """Запуск планировщика при старте приложения"""
    global scheduler
    if schedule_config["enabled"]:
        scheduler = create_scheduler()
        day_names = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
        trigger = CronTrigger(
            day_of_week=day_names[schedule_config["day_of_week"]],
            hour=schedule_config["hour"],
            minute=schedule_config["minute"]
        )
        scheduler.add_job(scheduled_report, trigger)
        scheduler.start()
        print("Планировщик запущен:", schedule_config)

@app.post("/schedule")
async def set_schedule(
        enabled: bool = Form(False),
        day: int = Form(0),
        hour: int = Form(10),
        minute: int = Form(0)
):
    """Установка расписания"""
    global scheduler, schedule_config

    # Полностью останавливаем старый планировщик
    if scheduler is not None:
        if scheduler.running:
            scheduler.shutdown(wait=True)  # Ждем завершения
        scheduler = None

    # Обновляем конфигурацию
    schedule_config.update({
        "enabled": enabled,
        "day_of_week": day,
        "hour": hour,
        "minute": minute
    })

    day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

    if enabled:
        # Создаем НОВЫЙ планировщик
        scheduler = create_scheduler()
        day_names_cron = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
        trigger = CronTrigger(
            day_of_week=day_names_cron[day],
            hour=hour,
            minute=minute
        )
        scheduler.add_job(scheduled_report, trigger)
        scheduler.start()
        print(f"Запланировано: каждый {day_names[day]} {hour:02d}:{minute:02d}")
        return {"status": "ok", "message": f"Запланировано: каждый {day_names[day]} {hour:02d}:{minute:02d}"}

    print("Планировщик остановлен")
    return {"status": "ok", "message": "Планировщик остановлен"}

@app.get("/status")
async def status():
    """Статус планировщика и файла"""
    global scheduler
    jobs_count = len(scheduler.get_jobs()) if scheduler and scheduler.running else 0
    file_size = os.path.getsize(settings.excel_path) if os.path.exists(settings.excel_path) else 0
    file_date = datetime.fromtimestamp(os.path.getmtime(settings.excel_path)).strftime("%d.%m %H:%M") if os.path.exists(
        settings.excel_path) else "Нет файла"

    day_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    next_run = f"{day_names[schedule_config['day_of_week']]} {schedule_config['hour']:02d}:{schedule_config['minute']:02d}" if \
    schedule_config['enabled'] else "Выключен"

    return {
        "enabled": schedule_config["enabled"],
        "scheduler_running": scheduler.running if scheduler else False,
        "next_run": next_run,
        "jobs_count": jobs_count,
        "file_size": file_size,
        "file_date": file_date,
        "last_dates": get_last_week_dates()
    }


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Главная страница с настройками расписания"""
    return templates.TemplateResponse("schedule.html", {
        "request": request,
        "config": schedule_config
    })


@app.post("/start-report", response_model=dict)
async def start_report(background_tasks: BackgroundTasks):
    """Запуск ручной выгрузки (для API)"""
    background_tasks.add_task(run_miac_direct)
    task_id = f"direct_{int(time.time())}"
    return {"task_id": task_id, "status": "started", "time": datetime.now().strftime("%H:%M:%S")}


@app.get("/manual")
async def manual_report(background_tasks: BackgroundTasks):
    """Ручной запуск с главной страницы"""
    background_tasks.add_task(run_miac_direct)
    return {"status": "started", "time": datetime.now().strftime("%H:%M:%S")}


@app.get("/download-report")
async def download_report():
    """Скачать готовый отчет"""
    if not os.path.exists(settings.excel_path):
        raise HTTPException(404, "Отчет не готов. Запустите выгрузку!")
    return FileResponse(
        settings.excel_path,
        filename="миац_отчет.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

EXTENSION_DIR = Path(__file__).resolve().parent / "extension"

@app.get("/download-extension")
async def download_extension():
    """Скачать расширение (zip с содержимым папки extension)"""
    if not EXTENSION_DIR.exists():
        raise HTTPException(404, "Папка extension не найдена на сервере")

    # Можно сделать уникальный файл, чтобы не было гонок при одновременных скачиваниях
    tmp_zip = Path(tempfile.gettempdir()) / f"Dashbord_extension_{int(time.time())}.zip"

    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in EXTENSION_DIR.rglob("*"):
            if p.is_file():
                # В архиве будут файлы в корне (manifest.json, sd.js, img/...)
                z.write(p, arcname=p.relative_to(EXTENSION_DIR))

    return FileResponse(
        path=str(tmp_zip),
        filename="Dashbord-extension.zip",
        media_type="application/zip",
    )


@app.get("/last-week-dates")
async def dates():
    """Даты прошлой недели (для отладки)"""
    return {"from": get_last_week_dates()[0], "to": get_last_week_dates()[1]}


@app.get("/docs")
async def docs_redirect():
    """Перенаправление на Swagger"""
    return {"docs": "/docs", "ui": "http://localhost:8000/docs"}